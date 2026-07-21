from flask import Flask, render_template, request, redirect, url_for, jsonify, send_file, make_response, session, g, flash
import sqlite3
import sys
import json
import os
import re 
import datetime as dt
import webview
import pandas as pd
from waitress import serve
import threading
from werkzeug.utils import secure_filename
import atexit
import io
import xlsxwriter
import base64
from pathlib import Path
from datetime import datetime, date

# --- OPENPYXL ENTEGRASYONU (HZM CORE) ---
import openpyxl
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.drawing.image import Image as OpenpyxlImage
# ----------------------------------------

# =================================================================
# DOSYA YOLLARI VE YAPILANDIRMA (EXE VE PY UYUMLU)
# =================================================================
if getattr(sys, 'frozen', False):
    bundle_dir = sys._MEIPASS
    BASE_DIR = os.path.dirname(sys.executable)
    template_folder = os.path.join(bundle_dir, 'templates')
    static_folder = os.path.join(bundle_dir, 'static')
else:
    BASE_DIR = os.path.abspath(os.path.dirname(__file__))
    template_folder = os.path.join(BASE_DIR, 'templates')
    static_folder = os.path.join(BASE_DIR, 'static')

app = Flask(__name__, template_folder=template_folder, static_folder=static_folder)
app.secret_key = "hzm_nexus_control_tsi_secure"
app.config['TEMPLATES_AUTO_RELOAD'] = True

UPLOAD_FOLDER = os.path.join(BASE_DIR, 'static', 'uploads', 'logos')
os.makedirs(UPLOAD_FOLDER, exist_ok=True)

# =================================================================
# AKILLI VERİTABANI ONARMA SİSTEMİ (MAIN.DB KURTARMA)
# =================================================================
def repair_main_db():
    main_db_path = os.path.join(BASE_DIR, "main.db")
    conn = sqlite3.connect(main_db_path)
    conn.execute("""CREATE TABLE IF NOT EXISTS companies (
                    id INTEGER PRIMARY KEY AUTOINCREMENT, 
                    name TEXT, 
                    phone TEXT, 
                    address TEXT, 
                    logo TEXT, 
                    tax_no TEXT, 
                    is_deleted INTEGER DEFAULT 0)""")
    conn.commit()

    cursor = conn.cursor()
    cursor.execute("SELECT COUNT(*) FROM companies WHERE is_deleted = 0")
    
    if cursor.fetchone()[0] == 0:
        print("Sistem Uyarısı: main.db boş veya kayıp. Şirket dosyaları taranıyor...")
        files = [f for f in os.listdir(BASE_DIR) if re.match(r'company_(\d+)\.db', f)]
        
        for file_name in files:
            try:
                c_id = re.findall(r'\d+', file_name)[0]
                c_path = os.path.join(BASE_DIR, file_name)
                
                c_name = f"Şirket {c_id}"
                c_is_deleted = 0
                c_phone = ""
                c_address = ""
                c_tax = ""

                if os.path.exists(c_path):
                    with sqlite3.connect(c_path) as c_conn:
                        c_cur = c_conn.cursor()
                        c_cur.execute("SELECT name FROM sqlite_master WHERE type='table' AND name='settings'")
                        if c_cur.fetchone():
                            settings = dict(c_cur.execute("SELECT key, value FROM settings").fetchall())
                            c_name = settings.get('company_name', c_name)
                            c_is_deleted = int(settings.get('is_deleted', 0))
                            c_phone = settings.get('phone', "")
                            c_address = settings.get('address', "")
                            c_tax = settings.get('tax_no', "")

                if c_is_deleted == 1:
                    continue

                conn.execute("""INSERT INTO companies (id, name, phone, address, logo, tax_no, is_deleted) 
                               VALUES (?, ?, ?, ?, '', ?, 0)""", 
                               (c_id, c_name, c_phone, c_address, c_tax))
                print(f"Başarıyla Onarıldı: {c_name} (ID: {c_id})")
                
            except Exception as e:
                print(f"Hata: {file_name} taranırken sorun çıktı: {e}")
        
        conn.commit()
    conn.close()

def check_main_db_schema():
    conn = sqlite3.connect(os.path.join(BASE_DIR, "main.db"))
    try:
        conn.execute("""
            CREATE TABLE IF NOT EXISTS companies (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                name TEXT NOT NULL,
                created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
                is_deleted INTEGER DEFAULT 0
            );
        """)
        conn.execute("""
            CREATE TABLE IF NOT EXISTS notes (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                employee_id INTEGER NULL,
                date TEXT NOT NULL,
                title TEXT NULL,
                content TEXT NOT NULL,
                created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
                FOREIGN KEY(employee_id) REFERENCES employees(id) ON DELETE CASCADE
            );
        """)
        conn.commit()
    except Exception as e:
        print(f"Şema Kontrol Motoru Hatası: {e}")
    finally:
        conn.close()

# =================================================================
# VERİTABANI YARDIMCILARI
# =================================================================

def get_conn():
    company_id = session.get('company_id')
    if not company_id: return None
    
    db_path = os.path.join(BASE_DIR, f"company_{company_id}.db")
    try:
        conn = sqlite3.connect(db_path, timeout=30)
        conn.row_factory = sqlite3.Row
        cursor = conn.cursor()
        
        cursor.execute("PRAGMA table_info(employees)")
        emp_cols = [column[1] for column in cursor.fetchall()]
        
        needed_emp_cols = {
            'phone': 'TEXT',
            'tc_no': 'TEXT',
            'address': 'TEXT',
            'entry_date': 'TEXT',
            'min_pay': 'REAL DEFAULT 0',
            'daily_wage': 'REAL DEFAULT 0',
            'company_id': 'INTEGER'
        }
        
        for col, col_type in needed_emp_cols.items():
            if col not in emp_cols:
                cursor.execute(f"ALTER TABLE employees ADD COLUMN {col} {col_type}")

        cursor.execute("PRAGMA table_info(attendance)")
        att_cols = [column[1] for column in cursor.fetchall()]
        
        needed_att_cols = {
            'advance': 'REAL DEFAULT 0',
            'daily_wage': 'REAL DEFAULT 0',
            'work_value': 'REAL DEFAULT 0',
            'extra_hours': 'REAL DEFAULT 0'
        }
        
        for col, col_type in needed_att_cols.items():
            if col not in att_cols:
                cursor.execute(f"ALTER TABLE attendance ADD COLUMN {col} {col_type}")
        
        conn.commit()
        return conn

    except Exception as e:
        print(f"HZM KRİTİK VERİTABANI HATASI: {e}")
        return None

def db_auto_fix():
    db_path = os.path.join(BASE_DIR, 'main.db')
    conn = sqlite3.connect(db_path)
    cursor = conn.cursor()
    try:
        cursor.execute("""
            CREATE TABLE IF NOT EXISTS employees (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                name TEXT,
                daily_wage REAL,
                phone TEXT,
                tc_no TEXT,
                entry_date TEXT,
                address TEXT
            )
        """)
        cursor.execute("PRAGMA table_info(employees)")
        columns = [column[1] for column in cursor.fetchall()]
        
        if 'min_pay' not in columns:
            cursor.execute("ALTER TABLE employees ADD COLUMN min_pay REAL DEFAULT 0")
            
        cursor.execute("""
            CREATE TABLE IF NOT EXISTS wage_history (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                employee_id INTEGER,
                wage REAL,
                valid_from TEXT
            )
        """)
        conn.commit()
    except Exception as e:
        print(f"HZM SİSTEMİ KRİTİK HATA: {e}")
    finally:
        conn.close()

db_auto_fix()

def init_main_db():
    main_db_path = os.path.join(BASE_DIR, "main.db")
    try:
        repair_main_db()
        with sqlite3.connect(main_db_path, timeout=30) as main_db:
            main_db.row_factory = sqlite3.Row
            columns = [
                ("phone", "TEXT"),
                ("address", "TEXT"),
                ("logo", "TEXT"),
                ("tax_no", "TEXT"),
                ("is_deleted", "INTEGER DEFAULT 0")
            ]
            for col_name, col_type in columns:
                try:
                    main_db.execute(f"ALTER TABLE companies ADD COLUMN {col_name} {col_type}")
                except sqlite3.OperationalError:
                    pass 
            main_db.commit()

        files = [f for f in os.listdir(BASE_DIR) if f.startswith("company_") and f.endswith(".db")]
        for f in files:
            db_path = os.path.join(BASE_DIR, f)
            try:
                with sqlite3.connect(db_path) as conn:
                    conn.row_factory = sqlite3.Row
                    cursor = conn.cursor()
                    cursor.executescript("""
                        CREATE TABLE IF NOT EXISTS employees (
                            id INTEGER PRIMARY KEY AUTOINCREMENT,
                            name TEXT NOT NULL,
                            daily_wage REAL DEFAULT 0,
                            company_id INTEGER
                        );
                        CREATE TABLE IF NOT EXISTS attendance (
                            id INTEGER PRIMARY KEY AUTOINCREMENT,
                            employee_id INTEGER,
                            date TEXT,
                            work_value REAL DEFAULT 0,
                            daily_wage REAL DEFAULT 0,
                            extra_hours REAL DEFAULT 0,
                            UNIQUE(employee_id, date)
                        );
                        CREATE TABLE IF NOT EXISTS payments (
                            id INTEGER PRIMARY KEY AUTOINCREMENT,
                            employee_id INTEGER,
                            amount REAL,
                            date TEXT,
                            description TEXT
                        );
                        CREATE TABLE IF NOT EXISTS notes (
                            id INTEGER PRIMARY KEY AUTOINCREMENT,
                            employee_id INTEGER NULL,
                            date TEXT,
                            title TEXT,
                            content TEXT,
                            created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
                        );
                    """)
                    conn.commit()
            except Exception as inner_e:
                print(f"Uyarı: {f} dosyası güncellenemedi: {inner_e}")

    except Exception as e:
        print(f"Kritik Hata: Ana sistem başlatılamadı: {e}")

init_main_db()

def get_unique_filename(path):
    if not os.path.exists(path):
        return path
    filename, extension = os.path.splitext(path)
    counter = 1
    while os.path.exists(f"{filename} ({counter}){extension}"):
        counter += 1
    return f"{filename} ({counter}){extension}"

# =================================================================
# MERKEZİ KONTROLLER
# =================================================================
@app.before_request
def global_control():
    if request.path.startswith('/static'): return
    g.active_month = session.get('active_month', dt.datetime.now().strftime('%Y-%m'))
    
    allowed_endpoints = [
        'select_company', 'add_company', 'select_company_set', 'static', 
        'get_deleted_companies', 'restore_company', 'delete_company', 
        'api_verify_admin', 'auto_fix_databases'
    ]
    
    if 'company_id' not in session and request.endpoint not in allowed_endpoints:
        return redirect(url_for('select_company'))

@app.context_processor
def inject_global_vars():
    return {
        'active_month': session.get('active_month', dt.datetime.now().strftime('%Y-%m')),
        'session': session,
        'now': dt.datetime.now().strftime('%d.%m.%Y %H:%M'),
        'datetime': dt.datetime
    }

# =================================================================
# ŞİRKET YÖNETİMİ & OTOMATİK KURTARMA
# =================================================================
@app.route("/select-company")
def select_company():
    main_db_path = os.path.join(BASE_DIR, "main.db")
    if not os.path.exists(main_db_path): init_main_db()
    
    try:
        with sqlite3.connect(main_db_path, timeout=30) as conn:
            conn.row_factory = sqlite3.Row
            companies = conn.execute("SELECT * FROM companies WHERE is_deleted = 0").fetchall()
        return render_template("select_company.html", companies=companies, has_pass='false')
    except Exception as e:
        return f"Ana veritabanına erişilemiyor: {e}", 500

@app.route("/auto-fix")
def auto_fix_databases():
    files = [f for f in os.listdir(BASE_DIR) if f.startswith("company_") and f.endswith(".db")]
    recovered = []
    with sqlite3.connect(os.path.join(BASE_DIR, "main.db"), timeout=30) as conn:
        for f in files:
            try:
                c_id = int(f.split("_")[1].split(".")[0])
                c_path = os.path.join(BASE_DIR, f)
                
                is_deleted_in_file = 0
                c_name = f"Kurtarılan Şirket {c_id}"
                
                with sqlite3.connect(c_path) as c_conn:
                    c_cur = c_conn.cursor()
                    c_cur.execute("SELECT name FROM sqlite_master WHERE type='table' AND name='settings'")
                    if c_cur.fetchone():
                        set_dict = dict(c_cur.execute("SELECT key, value FROM settings").fetchall())
                        is_deleted_in_file = int(set_dict.get('is_deleted', 0))
                        c_name = set_dict.get('company_name', c_name)
                
                if is_deleted_in_file == 1:
                    continue

                exists = conn.execute("SELECT id FROM companies WHERE id = ?", (c_id,)).fetchone()
                if not exists:
                    conn.execute("""INSERT INTO companies (id, name, is_deleted) 
                                   VALUES (?, ?, 0)""", (c_id, c_name))
                    recovered.append(c_name)
            except: continue
        conn.commit()
    return jsonify({"success": True, "recovered": recovered})

@app.route("/set-company/<int:id>", methods=["GET", "POST"])
def select_company_set(id):
    session.clear()
    with sqlite3.connect(os.path.join(BASE_DIR, "main.db")) as conn:
        conn.row_factory = sqlite3.Row
        company = conn.execute("SELECT name FROM companies WHERE id = ?", (id,)).fetchone()
    
    session['company_id'] = id
    session.permanent = True
    if company:
        session['company_name'] = company['name']
    
    if request.method == "POST":
        return jsonify({"success": True})
    return redirect(url_for('index'))

@app.route("/add-company", methods=["POST"])
def add_company():
    name = request.form.get("name")
    phone = request.form.get("phone")
    address = request.form.get("address")
    logo = request.files.get("logo")
    
    logo_filename = ""
    if logo and logo.filename != '':
        filename = secure_filename(f"{dt.datetime.now().timestamp()}_{logo.filename}")
        logo.save(os.path.join(UPLOAD_FOLDER, filename))
        logo_filename = f"static/uploads/logos/{filename}"
    
    with sqlite3.connect(os.path.join(BASE_DIR, "main.db")) as conn:
        cursor = conn.execute("INSERT INTO companies (name, phone, address, logo, is_deleted) VALUES (?, ?, ?, ?, 0)",
                             (name, phone, address, logo_filename))
        new_id = cursor.lastrowid
        conn.commit()

    db_path = os.path.join(BASE_DIR, f"company_{new_id}.db")
    with sqlite3.connect(db_path) as c_conn:
        c_conn.execute("PRAGMA journal_mode=WAL")
        c_conn.executescript("""
            CREATE TABLE IF NOT EXISTS settings (key TEXT PRIMARY KEY, value TEXT);
            CREATE TABLE IF NOT EXISTS employees (
                id INTEGER PRIMARY KEY AUTOINCREMENT, name TEXT, daily_wage REAL, 
                min_pay REAL DEFAULT 0, phone TEXT, tc_no TEXT, entry_date TEXT, address TEXT
            );
            CREATE TABLE IF NOT EXISTS wage_history (
                id INTEGER PRIMARY KEY AUTOINCREMENT, employee_id INTEGER, 
                wage REAL, valid_from TEXT
            );
            CREATE TABLE IF NOT EXISTS attendance (
                id INTEGER PRIMARY KEY AUTOINCREMENT, employee_id INTEGER, 
                date TEXT, work_value REAL DEFAULT 0, advance REAL DEFAULT 0, 
                daily_wage REAL DEFAULT 0, extra_hours REAL DEFAULT 0, UNIQUE(employee_id, date)
            );
            CREATE TABLE IF NOT EXISTS payments (
                id INTEGER PRIMARY KEY AUTOINCREMENT, employee_id INTEGER, 
                amount REAL, date TEXT, description TEXT
            );
            CREATE TABLE IF NOT EXISTS notes (
                id INTEGER PRIMARY KEY AUTOINCREMENT, employee_id INTEGER NULL, 
                date TEXT, title TEXT, content TEXT, created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
            );
        """)
        c_conn.execute("INSERT OR REPLACE INTO settings (key, value) VALUES ('company_name', ?)", (name,))
        c_conn.commit()

    return redirect(url_for('select_company'))

@app.route("/get-deleted-companies")
def get_deleted_companies():
    try:
        with sqlite3.connect(os.path.join(BASE_DIR, "main.db"), timeout=30) as conn:
            conn.row_factory = sqlite3.Row
            rows = conn.execute("SELECT id, name FROM companies WHERE is_deleted = 1").fetchall()
            return jsonify([dict(r) for r in rows])
    except:
        return jsonify([])

@app.route("/restore-company/<int:id>", methods=["POST"])
def restore_company(id):
    try:
        with sqlite3.connect(os.path.join(BASE_DIR, "main.db"), timeout=30) as conn:
            conn.execute("UPDATE companies SET is_deleted = 0 WHERE id = ?", (id,))
            conn.commit()
            
        c_path = os.path.join(BASE_DIR, f"company_{id}.db")
        if os.path.exists(c_path):
            with sqlite3.connect(c_path) as c_conn:
                c_conn.execute("INSERT OR REPLACE INTO settings (key, value) VALUES ('is_deleted', '0')")
                c_conn.commit()
                
        return jsonify({"success": True})
    except Exception as e:
        return jsonify({"success": False, "message": str(e)})

@app.route("/delete-company/<int:id>", methods=["POST"])
def delete_company(id):
    try:
        with sqlite3.connect(os.path.join(BASE_DIR, "main.db")) as conn:
            conn.execute("UPDATE companies SET is_deleted = 1 WHERE id = ?", (id,))
            conn.commit()
        
        c_path = os.path.join(BASE_DIR, f"company_{id}.db")
        if os.path.exists(c_path):
            with sqlite3.connect(c_path) as c_conn:
                c_conn.execute("CREATE TABLE IF NOT EXISTS settings (key TEXT PRIMARY KEY, value TEXT)")
                c_conn.execute("INSERT OR REPLACE INTO settings (key, value) VALUES ('is_deleted', '1')")
                c_conn.commit()

        if session.get('company_id') == id: session.clear()
        return jsonify({"success": True})
    except Exception as e:
        return jsonify({"success": False, "message": str(e)})

@app.route("/api/verify-admin", methods=["POST"])
def api_verify_admin():
    return jsonify({"success": True})

# =================================================================
# ANA PANEL VE PERSONEL MODÜLLERİ
# =================================================================
@app.route("/")
def index():
    return render_template("index.html")

@app.route("/patron-ekrani")
def patron_ekrani():
    conn = get_conn()
    if not conn: 
        return redirect(url_for('select_company'))
        
    conn.row_factory = sqlite3.Row
    cursor = conn.cursor()
    
    target_date = request.args.get('target_date', '').strip()
    if not target_date or target_date == "None":
        target_date = datetime.now().strftime('%Y-%m-%d')
        
    try:
        cursor.execute("""
            SELECT COALESCE(SUM((work_value + (COALESCE(extra_hours, 0) / 8.0)) * daily_wage), 0) as total_brut
            FROM attendance WHERE daily_wage > 0
        """)
        sirket_toplam_brut = float(cursor.fetchone()['total_brut'] or 0)
        
        cursor.execute("SELECT COALESCE(SUM(amount), 0) as total_paid FROM payments")
        sirket_toplam_odenen = float(cursor.fetchone()['total_paid'] or 0)
        
        sirket_net_borc = sirket_toplam_brut - sirket_toplam_odenen

        cursor.execute("""
            SELECT COALESCE(SUM((work_value + (COALESCE(extra_hours, 0) / 8.0)) * daily_wage), 0) as daily_brut
            FROM attendance WHERE date = ? AND daily_wage > 0
        """, (target_date,))
        daily_total = float(cursor.fetchone()['daily_brut'] or 0)
        
        cursor.execute("SELECT COALESCE(SUM(amount), 0) as daily_paid FROM payments WHERE date = ?", (target_date,))
        daily_avans = float(cursor.fetchone()['daily_paid'] or 0)

        all_staff_rows = cursor.execute("SELECT id, name FROM employees ORDER BY name ASC").fetchall()
        active_staff = len(all_staff_rows)
        
        top_overtime = cursor.execute("""
            SELECT e.name, COALESCE(SUM(a.extra_hours), 0) as toplam_saat
            FROM employees e
            JOIN attendance a ON e.id = a.employee_id
            GROUP BY e.id
            HAVING toplam_saat > 0
            ORDER BY toplam_saat DESC
            LIMIT 5
        """).fetchall()

        cursor.execute("""
            SELECT e.name, 'Puantaj Girişi' as islem, ((a.work_value + (COALESCE(a.extra_hours, 0)/8.0)) * a.daily_wage) as tutar
            FROM attendance a
            JOIN employees e ON a.employee_id = e.id
            WHERE a.date = ? AND (a.work_value > 0 OR a.extra_hours > 0)
        """, (target_date,))
        p_logs = cursor.fetchall()
        
        cursor.execute("""
            SELECT e.name, p.description as islem, p.amount as tutar
            FROM payments p
            JOIN employees e ON p.employee_id = e.id
            WHERE p.date = ?
        """, (target_date,))
        pay_logs = cursor.fetchall()
        
        transactions = []
        for row in p_logs:
            transactions.append({"name": row['name'], "type": row['islem'], "amount": float(row['tutar'] or 0)})
        for row in pay_logs:
            transactions.append({"name": row['name'], "type": row['islem'], "amount": float(row['tutar'] or 0)})
            
        transactions.sort(key=lambda x: x['amount'], reverse=True)

        notes_rows = cursor.execute("""
            SELECT n.id, n.title, n.content, n.employee_id, e.name as emp_name, n.created_at
            FROM notes n
            LEFT JOIN employees e ON n.employee_id = e.id
            WHERE n.date = ?
            ORDER BY n.id DESC
        """, (target_date,)).fetchall()

        return render_template(
            "patron.html",
            selected_date=target_date,
            daily_total=daily_total,
            daily_avans=daily_avans,
            active_staff=active_staff,
            sirket_toplam_brut=sirket_toplam_brut,
            sirket_toplam_odenen=sirket_toplam_odenen,
            sirket_net_borc=sirket_net_borc,
            top_overtime=top_overtime,
            transactions=transactions,
            all_staff=all_staff_rows,
            notes_list=notes_rows
        )
        
    except Exception as e:
        print(f"TSİ Patron Paneli Kritik Hata: {e}")
        return f"Sistem Hatası: {e}", 500
    finally:
        conn.close()

# =================================================================
# AKILLI HAFIZA NOT SİSTEMİ API KÖPRÜLERİ
# =================================================================
@app.route("/save-note", methods=["POST"])
def save_note():
    conn = get_conn()
    if not conn:
        return jsonify({"success": False, "message": "Bağlantı kesildi"}), 400
        
    try:
        note_id = request.form.get("note_id", "").strip()
        title = request.form.get("title", "").strip()
        content = request.form.get("content", "").strip()
        note_date = request.form.get("date", "").strip()
        emp_id = request.form.get("employee_id", "").strip()
        
        if not content or not note_date:
            return jsonify({"success": False, "message": "Not içeriği veya tarih eksik"}), 400
            
        if not title:
            words = content.split()
            if len(words) >= 2:
                title = f"{words[0]} {words[1]}"
            elif len(words) == 1:
                title = words[0]
            else:
                title = "Adsız Not"
        
        db_emp_id = int(emp_id) if (emp_id and emp_id != "all") else None
        
        if note_id:
            conn.execute("""
                UPDATE notes 
                SET employee_id = ?, title = ?, content = ? 
                WHERE id = ?
            """, (db_emp_id, title, content, int(note_id)))
            message = "Hafıza kaydı güncellendi."
        else:
            conn.execute("""
                INSERT INTO notes (employee_id, date, title, content)
                VALUES (?, ?, ?, ?)
            """, (db_emp_id, note_date, title, content))
            message = "Hafıza kaydı mühürlendi."
            
        conn.commit()
        return jsonify({"success": True, "message": message})
    except Exception as e:
        print(f"Not Kayıt Hatası: {e}")
        return jsonify({"success": False, "message": str(e)}), 500
    finally:
        conn.close()

@app.route("/delete-note/<int:id>", methods=["POST"])
def delete_note(id):
    conn = get_conn()
    if not conn:
        return jsonify({"success": False, "message": "Bağlantı kesildi"}), 400
    try:
        conn.execute("DELETE FROM notes WHERE id = ?", (id,))
        conn.commit()
        return jsonify({"success": True, "message": "Not hafızadan silindi."})
    except Exception as e:
        return jsonify({"success": False, "message": str(e)}), 500
    finally:
        conn.close()

@app.route("/calisanlar", methods=["GET", "POST"])
def calisanlar():
    conn = get_conn()
    if not conn: return redirect(url_for('select_company'))
    
    if request.method == "POST":
        try:
            name = request.form.get('name')
            wage = request.form.get('daily_wage', '0').replace(',', '.')
            min_pay = request.form.get('min_pay', '0').replace(',', '.') 
            phone = request.form.get('phone', '')
            tc = request.form.get('tc_no', '')
            entry = request.form.get('entry_date', '')
            address = request.form.get('address', '')
            
            conn.execute("""
                INSERT INTO employees (name, daily_wage, min_pay, phone, tc_no, entry_date, address) 
                VALUES (?, ?, ?, ?, ?, ?, ?)""", 
                (name, float(wage), float(min_pay), phone, tc, entry, address))
            conn.commit()
        except Exception as e:
            print(f"TSİ Personel Ekleme Hatası: {e}")
        finally: 
            conn.close()
        return redirect(url_for('calisanlar'))
    
    try:
        conn.row_factory = lambda cursor, row: dict((cursor.description[i][0], item) for i, item in enumerate(row))
        emps = conn.execute("SELECT * FROM employees ORDER BY id DESC").fetchall()
        now_str = datetime.now().strftime('%Y-%m-%d')
        return render_template("calisanlar.html", employees=emps, now=now_str)
    finally: 
        conn.close()

@app.route('/api/personel-detay-full/<int:id>')
def personel_detay_api(id):
    conn = None
    try:
        conn = get_conn()
        conn.row_factory = lambda cursor, row: dict((cursor.description[i][0], item) for i, item in enumerate(row))
        cursor = conn.cursor()
        
        cursor.execute("SELECT * FROM employees WHERE id = ?", (id,))
        emp_row = cursor.fetchone()
        emp_info = dict(emp_row) if emp_row else {}

        query = """
            SELECT 
                date, 
                work_value, 
                COALESCE(extra_hours, 0) as raw_extra_hours, 
                (COALESCE(extra_hours, 0) / 8.0) as overtime_hours, 
                daily_wage, 
                advance, 
                'Puantaj' as desc
            FROM attendance 
            WHERE employee_id = ? AND (work_value > 0 OR extra_hours > 0)
            
            UNION ALL
            
            SELECT 
                date, 
                0 as work_value, 
                0 as raw_extra_hours,
                0 as overtime_hours, 
                0 as daily_wage, 
                amount as advance, 
                description as desc
            FROM payments 
            WHERE employee_id = ?
            
            ORDER BY date ASC
        """
        cursor.execute(query, (id, id))
        history = cursor.fetchall()
        
        for h in history:
            try:
                w_val = float(h.get('work_value') or 0)
                ot_days = float(h.get('overtime_hours') or 0) 
                d_wage = float(h.get('daily_wage') or 0)
                
                h['hakedis'] = (w_val + ot_days) * d_wage
                
                h['work_value'] = w_val
                h['overtime_hours'] = ot_days
                h['daily_wage'] = d_wage
            except (ValueError, TypeError):
                h['hakedis'] = 0.0
        
        return jsonify({"info": emp_info, "history": history})
        
    except Exception as e:
        print(f"TSİ KRİTİK HATA: {e}")
        return jsonify({"error": str(e), "info": {}, "history": []}), 500
        
    finally:
        if conn:
            conn.close()

@app.route('/calisan-guncelle/<int:id>', methods=['POST'])
def calisan_guncelle(id):
    conn = get_conn()
    cursor = conn.cursor()
    
    try:
        name = request.form.get('name')
        tc_no = request.form.get('tc_no')
        phone = request.form.get('phone')
        new_wage = float(request.form.get('daily_wage', '0').replace(',', '.'))
        
        start_date = request.form.get('wage_start_date') 
        end_date = request.form.get('wage_end_date')     
        
        cursor.execute("""
            UPDATE employees SET name=?, tc_no=?, phone=?, daily_wage=? 
            WHERE id=?
        """, (name, tc_no, phone, new_wage, id))

        if start_date and end_date and new_wage > 0:
            cursor.execute("""
                UPDATE attendance 
                SET daily_wage = ? 
                WHERE employee_id = ? AND date BETWEEN ? AND ?
            """, (new_wage, id, start_date, end_date))

        conn.commit()
        flash("Belirtilen tarih aralığındaki yevmiyeler mühürlendi.", "success")
    except Exception as e:
        print(f"TSİ Güncelleme Hatası: {e}")
    finally:
        conn.close()
        
    return redirect(url_for('calisanlar'))

@app.route("/calisan-sil/<int:id>")
def calisan_sil(id):
    conn = get_conn()
    if conn:
        try:
            conn.execute("DELETE FROM employees WHERE id = ?", (id,))
            conn.execute("DELETE FROM attendance WHERE employee_id = ?", (id,))
            conn.execute("DELETE FROM payments WHERE employee_id = ?", (id,))
            conn.commit()
        except Exception as e:
            print(f"TSİ Silme Hatası: {e}")
        finally:
            conn.close()
    return redirect(url_for('calisanlar'))

# =================================================================
# PUANTAJ VE TOPLU İŞLEMLER
# =================================================================
@app.route("/puantaj")
def puantaj():
    month = request.args.get('month', g.active_month)
    conn = get_conn()
    if not conn: return redirect(url_for('select_company'))
    
    try:
        emps = conn.execute("SELECT * FROM employees ORDER BY name ASC").fetchall()
        attendance_rows = conn.execute("""
            SELECT employee_id, date, work_value, extra_hours 
            FROM attendance 
            WHERE strftime('%Y-%m', date) = ?
        """, (month,)).fetchall()
        
        attendance_map = {}
        for row in attendance_rows:
            e_id = str(row['employee_id'])
            if e_id not in attendance_map:
                attendance_map[e_id] = {'data': {}, 'extra': 0}
            
            attendance_map[e_id]['data'][row['date']] = row['work_value']
            if row['extra_hours'] and row['extra_hours'] > 0:
                attendance_map[e_id]['extra'] = row['extra_hours']
            
        return render_template("puantaj.html", employees=emps, month=month, attendance=attendance_map)
    finally:
        conn.close()

@app.route("/save_puantaj", methods=["POST"])
def save_puantaj():
    company_id = session.get('company_id')
    if not company_id:
        return jsonify({"success": False, "error": "Şirket seçili değil"}), 400
        
    req_data = request.get_json()
    if not req_data:
        return jsonify({"success": False, "error": "Veri alınamadı"}), 400

    month = req_data.get('month') 
    puantaj_verisi = req_data.get('puantaj', {})
    
    conn = get_conn()
    if not conn:
        return jsonify({"success": False, "error": "Veritabanı bağlantısı kurulamadı"}), 500
        
    cursor = conn.cursor()
    
    try:
        for emp_id, days in puantaj_verisi.items():
            for day, values in days.items():
                work_value = values.get('work', 0)
                extra_hours = values.get('extra', 0)
                advance_val = values.get('advance', 0)
                
                date_str = f"{month}-{str(day).zfill(2)}"
                
                cursor.execute("SELECT daily_wage FROM employees WHERE id = ?", (emp_id,))
                row = cursor.fetchone()
                daily_wage = row[0] if row and row[0] is not None else 0
                
                cursor.execute("""
                    SELECT id FROM attendance 
                    WHERE employee_id = ? AND date = ?
                """, (emp_id, date_str))
                
                exist = cursor.fetchone()
                if exist:
                    cursor.execute("""
                        UPDATE attendance 
                        SET work_value = ?, extra_hours = ?, daily_wage = ?, advance = ?
                        WHERE employee_id = ? AND date = ?
                    """, (work_value, extra_hours, daily_wage, advance_val, emp_id, date_str))
                else:
                    cursor.execute("""
                        INSERT INTO attendance (employee_id, date, work_value, extra_hours, daily_wage, advance)
                        VALUES (?, ?, ?, ?, ?, ?)
                    """, (emp_id, date_str, work_value, extra_hours, daily_wage, advance_val))
                    
        conn.commit()
        return jsonify({"success": True, "message": "Puantaj başarıyla mühürlendi."})
        
    except Exception as e:
        conn.rollback()
        print(f"Puantaj Kayıt Kritik Hatası: {e}")
        return jsonify({"success": False, "error": str(e)}), 500
        
    finally:
        conn.close()

@app.route("/toplu-giris", methods=["GET", "POST"])
def toplu_islem_paneli():
    conn = get_conn()
    if not conn: return redirect(url_for('select_company'))
    
    today = datetime.now()
    default_start = today.replace(day=1).strftime('%Y-%m-%d')
    default_end = today.strftime('%Y-%m-%d')
    
    start_date = request.args.get('start', default_start)
    end_date = request.args.get('end', default_end)
    today_date = today.strftime('%Y-%m-%d')
    
    try:
        if request.method == "POST":
            mode = request.form.get("mode")
            active_panel = "salary"
            
            if mode == "bulk_payment_entry":
                active_panel = "salary"
                emp_ids = request.form.getlist("emp_ids[]")
                pay1 = request.form.getlist("pay1[]")
                pay2 = request.form.getlist("pay2[]")
                pay3 = request.form.getlist("pay3[]")
                entry_date = request.form.get("entry_date") or today_date
                
                for i in range(len(emp_ids)):
                    configs = [(pay1, "Ödeme 1 (Maaş)"), (pay2, "Ödeme 2 (Prim)"), (pay3, "Ödeme 3 (Ekstra)")]
                    for p_list, desc in configs:
                        if i < len(p_list) and p_list[i]:
                            val = float(p_list[i].replace(',', '.'))
                            if val > 0:
                                conn.execute("INSERT INTO payments (employee_id, amount, date, description) VALUES (?, ?, ?, ?)",
                                             (emp_ids[i], val, entry_date, desc))
                conn.commit()

            elif mode == "advance_edit":
                active_panel = "manage"
                p_ids = request.form.getlist("payment_ids[]")
                dates = request.form.getlist("edit_dates[]")
                amounts = request.form.getlist("edit_amounts[]")
                descs = request.form.getlist("edit_descriptions[]")
                for pid, d, amt, ds in zip(p_ids, dates, amounts, descs):
                    if amt:
                        conn.execute("UPDATE payments SET date=?, amount=?, description=? WHERE id=?", 
                                     (d, float(str(amt).replace(',', '.')), ds, pid))
                conn.commit()

            return redirect(url_for('toplu_islem_paneli', start=start_date, end=end_date, active_panel=active_panel))

        employees = conn.execute("SELECT id, name FROM employees ORDER BY name ASC").fetchall()
        recent_advances = conn.execute("""
            SELECT p.id, e.name, p.amount, p.date, p.description 
            FROM payments p JOIN employees e ON p.employee_id = e.id 
            WHERE p.date BETWEEN ? AND ? 
            ORDER BY p.date DESC, p.id DESC
        """, (start_date, end_date)).fetchall()
        
        return render_template("toplu-giris.html", 
                               employees=employees, 
                               recent_advances=recent_advances, 
                               start_date=start_date, 
                               end_date=end_date, 
                               today_date=today_date)
    finally:
        conn.close()

@app.route("/avans-sil/<int:id>")
def avans_sil(id):
    conn = get_conn()
    month = request.args.get('month', g.active_month)
    if conn:
        conn.execute("DELETE FROM payments WHERE id = ?", (id,))
        conn.commit()
        conn.close()
    return redirect(url_for('toplu_islem_paneli', month=month))

# =================================================================
# RAPORLAMA VE ÇIKIŞ
# =================================================================
@app.route("/net-tablo")
def net_tablo():
    conn = get_conn()
    if not conn: return redirect(url_for('select_company'))
    conn.execute("PRAGMA journal_mode=WAL")
    conn.row_factory = sqlite3.Row
    
    start_date = request.args.get('start', datetime.now().replace(day=1).strftime('%Y-%m-%d'))
    end_date = request.args.get('end', datetime.now().strftime('%Y-%m-%d'))

    try:
        employees = conn.execute("SELECT id, name, daily_wage FROM employees ORDER BY name ASC").fetchall()
        
        # 1. GEÇMİŞTEN DEVİR HESAPLAMA (start_date ÖNCESİ TAM TARAMA)
        # A) Geçmiş Hak Edişler
        gp_rows = conn.execute("""
            SELECT a.employee_id, 
                   SUM((COALESCE(a.work_value, 0) + (COALESCE(a.extra_hours, 0) / 8.0)) * CASE WHEN a.daily_wage > 0 THEN a.daily_wage ELSE e.daily_wage END) as g_kazanc
            FROM attendance a 
            JOIN employees e ON a.employee_id = e.id
            WHERE date(a.date) < date(?) GROUP BY a.employee_id
        """, (start_date,)).fetchall()
        gecmis_kazanclar = {r['employee_id']: r['g_kazanc'] for r in gp_rows}

        # B) Geçmiş Asgari Maaş Ödemeleri
        gasgari_rows = conn.execute("""
            SELECT employee_id, SUM(amount) as g_asgari
            FROM payments WHERE date(date) < date(?) AND description = 'Ödeme 1 (Maaş)' GROUP BY employee_id
        """, (start_date,)).fetchall()
        gecmis_asgari = {r['employee_id']: r['g_asgari'] for r in gasgari_rows}

        # C) Geçmiş Avans Ödemeleri
        gavans_rows = conn.execute("""
            SELECT employee_id, SUM(amount) as g_avans
            FROM payments WHERE date(date) < date(?) AND (description IS NULL OR description != 'Ödeme 1 (Maaş)') GROUP BY employee_id
        """, (start_date,)).fetchall()
        gecmis_avans = {r['employee_id']: r['g_avans'] for r in gavans_rows}

        # 2. MEVCUT DÖNEM HAREKETLERİ (start_date ve end_date ARASI)
        mp_rows = conn.execute("""
            SELECT employee_id, SUM(work_value) as gun, SUM(extra_hours) as mesai, AVG(daily_wage) as ort_yevmiye
            FROM attendance WHERE date(date) BETWEEN date(?) AND date(?) GROUP BY employee_id
        """, (start_date, end_date)).fetchall()
        
        mo_rows = conn.execute("""
            SELECT employee_id, 
                   SUM(CASE WHEN description='Ödeme 1 (Maaş)' THEN amount ELSE 0 END) as asgari,
                   SUM(CASE WHEN description!='Ödeme 1 (Maaş)' THEN amount ELSE 0 END) as avans
            FROM payments WHERE date(date) BETWEEN date(?) AND date(?) GROUP BY employee_id
        """, (start_date, end_date)).fetchall()

        mevcut_puantaj = {r['employee_id']: r for r in mp_rows}
        mevcut_odeme = {r['employee_id']: r for r in mo_rows}

        # 3. DÖNEMSEL MİZAN VE DEVİR BİRLEŞTİRME
        hesap_list = []
        for emp in employees:
            eid = emp['id']
            
            # Devir = Geçmiş Kazanç - (Geçmiş Asgari + Geçmiş Avans)
            g_kazanc = gecmis_kazanclar.get(eid, 0) or 0
            g_asgari = gecmis_asgari.get(eid, 0) or 0
            g_avans = gecmis_avans.get(eid, 0) or 0
            devir = g_kazanc - g_asgari - g_avans
            
            p = mevcut_puantaj.get(eid)
            o = mevcut_odeme.get(eid)
            
            normal_gun = p['gun'] if p else 0
            mesai = p['mesai'] if p else 0
            yevmiye = p['ort_yevmiye'] if (p and p['ort_yevmiye']) else emp['daily_wage']
            
            brut = (normal_gun + (mesai / 8.0)) * yevmiye
            asgari = o['asgari'] if o else 0
            avans = o['avans'] if o else 0
            
            # Kalan Net = Devir Bakiyesi + Dönem Hak Ediş - Dönem Kesintiler
            net = devir + brut - asgari - avans
            
            if brut != 0 or devir != 0 or asgari != 0 or avans != 0:
                hesap_list.append({
                    "name": emp['name'], 
                    "devir": round(devir, 2), 
                    "normal_gun": normal_gun,
                    "mesai_saat": mesai, 
                    "mesai_gun": round(mesai / 8.0, 2), 
                    "ana_yevmiye": round(yevmiye, 2),
                    "yevmiye_kazanc": round(brut, 2), 
                    "asgari_tutar": asgari, 
                    "toplam_avans": avans, 
                    "net": round(net, 2)
                })

        return render_template("net-tablo.html", hesap_list=hesap_list, start_date=start_date, end_date=end_date)
    except Exception as e:
        print(f"Mizan Hatası: {e}")
        return f"Mizan Hatası: {e}"
    finally:
        conn.close()

@app.route("/finans-analiz")
def finans_analiz():
    return redirect(url_for('net_tablo', **request.args))

@app.route("/export-excel")
def export_excel():
    try:
        start_date = request.args.get('start', datetime.now().replace(day=1).strftime('%Y-%m-%d'))
        end_date = request.args.get('end', datetime.now().strftime('%Y-%m-%d'))
        
        # Mizan hesap motorunu doğrudan çalıştırarak hesap_list nesnesini alıyoruz
        conn = get_conn()
        if not conn: return "Veritabanı bağlantısı kurulamadı.", 500
        
        employees = conn.execute("SELECT id, name, daily_wage FROM employees ORDER BY name ASC").fetchall()
        hesap_list = []

        for emp in employees:
            eid = emp['id']
            devir_p = conn.execute("SELECT SUM((COALESCE(work_value,0) + COALESCE(extra_hours,0)/8.0) * COALESCE(daily_wage, ?)) FROM attendance WHERE employee_id=? AND date < ?", (emp['daily_wage'], eid, start_date)).fetchone()[0] or 0.0
            devir_asg = conn.execute("SELECT SUM(amount) FROM payments WHERE employee_id=? AND date < ? AND description='Ödeme 1 (Maaş)'", (eid, start_date)).fetchone()[0] or 0.0
            devir_avn = conn.execute("SELECT SUM(amount) FROM payments WHERE employee_id=? AND date < ? AND (description IS NULL OR description!='Ödeme 1 (Maaş)')", (eid, start_date)).fetchone()[0] or 0.0
            devir = devir_p - devir_asg - devir_avn

            p = conn.execute("SELECT SUM(work_value) as gun, SUM(extra_hours) as mesai, AVG(daily_wage) as ort_yevmiye FROM attendance WHERE employee_id=? AND date BETWEEN ? AND ?", (eid, start_date, end_date)).fetchone()
            o = conn.execute("SELECT SUM(CASE WHEN description='Ödeme 1 (Maaş)' THEN amount ELSE 0 END) as asgari, SUM(CASE WHEN description!='Ödeme 1 (Maaş)' THEN amount ELSE 0 END) as avans FROM payments WHERE employee_id=? AND date BETWEEN ? AND ?", (eid, start_date, end_date)).fetchone()

            normal_gun = p['gun'] or 0.0
            mesai = p['mesai'] or 0.0
            yevmiye = p['ort_yevmiye'] if (p and p['ort_yevmiye']) else (emp['daily_wage'] or 0.0)
            brut = (normal_gun + (mesai / 8.0)) * yevmiye
            asgari = o['asgari'] or 0.0
            avans = o['avans'] or 0.0
            net = devir + brut - asgari - avans

            if brut != 0 or devir != 0 or asgari != 0 or avans != 0:
                hesap_list.append({
                    "name": emp['name'], "devir": round(devir, 2), "normal_gun": normal_gun,
                    "mesai_saat": mesai, "mesai_gun": round(mesai / 8.0, 2), "ana_yevmiye": round(yevmiye, 2),
                    "yevmiye_kazanc": round(brut, 2), "asgari_tutar": asgari, "toplam_avans": avans, "net": round(net, 2)
                })
        conn.close()

        if not hesap_list:
            return "Aktarılacak mali veri bulunamadı.", 404

        base_name = f"TSI_Mali_Rapor_{start_date}.xlsx"
        initial_path = os.path.join(Path.home() / "Downloads", base_name)
        final_path = get_unique_filename(initial_path)
        
        writer = pd.ExcelWriter(final_path, engine='xlsxwriter')
        workbook  = writer.book
        worksheet = workbook.add_worksheet('Mali Rapor')

        title_fmt = workbook.add_format({'bold': True, 'font_size': 15, 'font_name': 'Arial', 'font_color': '#FFFFFF', 'bg_color': '#1A1A1A', 'align': 'center', 'valign': 'vcenter'})
        comp_fmt  = workbook.add_format({'bold': True, 'font_size': 11.5, 'font_name': 'Arial', 'font_color': '#0062FF', 'align': 'center', 'valign': 'vcenter'})
        date_fmt  = workbook.add_format({'italic': True, 'font_size': 9.5, 'font_name': 'Arial', 'font_color': '#555555', 'align': 'center', 'valign': 'vcenter'})
        
        header_fmt = workbook.add_format({'bold': True, 'bg_color': '#1A1A1A', 'font_color': '#FFFFFF', 'border': 1, 'border_color': '#333333', 'align': 'center', 'valign': 'vcenter'})
        cell_fmt   = workbook.add_format({'border': 1, 'border_color': '#E0E0E0', 'align': 'left'})
        center_fmt = workbook.add_format({'border': 1, 'border_color': '#E0E0E0', 'align': 'center'})
        
        num_fmt     = workbook.add_format({'border': 1, 'border_color': '#E0E0E0', 'align': 'right', 'num_format': '#,##0.00 ₺'})
        blue_fmt   = workbook.add_format({'font_color': '#0062FF', 'border': 1, 'border_color': '#E0E0E0', 'align': 'right', 'num_format': '- #,##0.00 ₺'})
        red_fmt     = workbook.add_format({'font_color': '#DC3545', 'border': 1, 'border_color': '#E0E0E0', 'align': 'right', 'num_format': '- #,##0.00 ₺'})
        green_bold = workbook.add_format({'font_color': '#28A745', 'bold': True, 'border': 1, 'border_color': '#E0E0E0', 'align': 'right', 'num_format': '#,##0.00 ₺'})
        red_bold    = workbook.add_format({'font_color': '#DC3545', 'bold': True, 'border': 1, 'border_color': '#E0E0E0', 'align': 'right', 'num_format': '#,##0.00 ₺'})

        total_label_fmt = workbook.add_format({'bold': True, 'bg_color': '#F2F2F2', 'border': 1, 'border_color': '#333333', 'align': 'left'})
        total_center_fmt = workbook.add_format({'bold': True, 'bg_color': '#F2F2F2', 'border': 1, 'border_color': '#333333', 'align': 'center'})
        total_num_fmt = workbook.add_format({'bold': True, 'bg_color': '#F2F2F2', 'border': 1, 'border_color': '#333333', 'align': 'right', 'num_format': '#,##0.00 ₺'})
        total_blue_fmt = workbook.add_format({'bold': True, 'bg_color': '#F2F2F2', 'font_color': '#0062FF', 'border': 1, 'border_color': '#333333', 'align': 'right', 'num_format': '- #,##0.00 ₺'})
        total_red_fmt = workbook.add_format({'bold': True, 'bg_color': '#F2F2F2', 'font_color': '#DC3545', 'border': 1, 'border_color': '#333333', 'align': 'right', 'num_format': '- #,##0.00 ₺'})
        total_net_fmt = workbook.add_format({'bold': True, 'bg_color': '#1A1A1A', 'font_color': '#28A745', 'border': 1, 'border_color': '#333333', 'align': 'right', 'num_format': '#,##0.00 ₺'})
        mühür_fmt = workbook.add_format({'italic': True, 'font_size': 9, 'font_name': 'Arial', 'font_color': '#333333', 'align': 'center', 'valign': 'vcenter'})

        worksheet.set_column('A:J', 18)

        worksheet.set_row(0, 32)  
        worksheet.set_row(1, 22)  
        worksheet.set_row(2, 20)  
        worksheet.set_row(4, 24)  

        worksheet.merge_range('A1:J1', 'TSİ INTEGRATION SYSTEM | MALİ RAPOR TABLOSU', title_fmt)
        worksheet.merge_range('A2:J2', 'TECHNICAL STATISTICS & INTEGRATION', comp_fmt)
        worksheet.merge_range('A3:J3', f'Rapor Dönemi: {start_date} - {end_date}', date_fmt)

        headers = ["PERSONEL", "GEÇMİŞ DEVİR", "TOPLAM GÜN", "MESAİ SAAT", "MESAİ GÜN", "G. YEVMİYE", "BRÜT TUTAR", "ASGARİ", "TOPLAM AVANS", "KALAN NET"]
        for col, head in enumerate(headers):
            worksheet.write(4, col, head, header_fmt)

        start_row = 5  
        current_row = start_row
          
        for item in hesap_list:
            worksheet.write(current_row, 0, str(item.get('name', '')).upper(), cell_fmt)
            worksheet.write(current_row, 1, item.get('devir', 0), num_fmt)
            worksheet.write(current_row, 2, item.get('normal_gun', 0), center_fmt)
            worksheet.write(current_row, 3, item.get('mesai_saat', 0), center_fmt)
            worksheet.write(current_row, 4, float("{:.2f}".format(item.get('mesai_gun', 0))), center_fmt)
            worksheet.write(current_row, 5, item.get('ana_yevmiye', 0), num_fmt)
            worksheet.write(current_row, 6, item.get('yevmiye_kazanc', 0), num_fmt)
            worksheet.write(current_row, 7, item.get('asgari_tutar', 0), blue_fmt)
            worksheet.write(current_row, 8, item.get('toplam_avans', 0), red_fmt)
            
            net_val = item.get('net', 0)
            worksheet.write(current_row, 9, net_val, green_bold if net_val >= 0 else red_bold)
            current_row += 1

        excel_start = start_row + 1
        excel_end = current_row

        worksheet.write(current_row, 0, "GENEL TOPLAM", total_label_fmt)
        worksheet.write_formula(current_row, 1, f"=SUM(B{excel_start}:B{excel_end})", total_num_fmt)
        worksheet.write_formula(current_row, 2, f"=SUM(C{excel_start}:C{excel_end})", total_center_fmt)
        worksheet.write_formula(current_row, 3, f"=SUM(D{excel_start}:D{excel_end})", total_center_fmt)
        worksheet.write_formula(current_row, 4, f"=SUM(E{excel_start}:E{excel_end})", total_center_fmt)
        worksheet.write(current_row, 5, "-", total_center_fmt) 
        worksheet.write_formula(current_row, 6, f"=SUM(G{excel_start}:G{excel_end})", total_num_fmt)
        worksheet.write_formula(current_row, 7, f"=SUM(H{excel_start}:H{excel_end})", total_blue_fmt)
        worksheet.write_formula(current_row, 8, f"=SUM(I{excel_start}:I{excel_end})", total_red_fmt)
        worksheet.write_formula(current_row, 9, f"=SUM(J{excel_start}:J{excel_end})", total_net_fmt)

        text_row = current_row + 2
        worksheet.set_row(text_row, 24)
        worksheet.merge_range(text_row, 0, text_row, 9, "Bu Belge HZM SOFTWARE & DESIGN SERVICES Tarafından Geliştirilen TSI: Technical Statistics & Integration Sistemi İle Hazırlanmıştır.", mühür_fmt)
        
        logo_row = text_row + 2
        worksheet.set_row(logo_row, 42) 
        
        logo_path = 'static/img/logo.png'
        logo_path2 = 'static/img/TSI_logo.png'
        compact_scale = 0.11  

        if os.path.exists(logo_path):
            worksheet.insert_image(logo_row, 0, logo_path, {'x_scale': compact_scale, 'y_scale': compact_scale, 'object_position': 1})
            
        if os.path.exists(logo_path2):
            worksheet.insert_image(logo_row, 9, logo_path2, {'x_scale': compact_scale, 'y_scale': compact_scale, 'object_position': 1})

        writer.close()
        return send_file(final_path, as_attachment=True)

    except Exception as e:
        print(f"TSİ Mali Rapor Çıktı Hatası: {e}")
        return f"Excel Oluşturma Hatası: {str(e)}", 500

@app.route("/export-puantaj-excel/<month>")
def export_puantaj_excel(month):
    try:
        conn = get_conn()
        if not conn: 
            return "Veritabanı bağlantısı kurulamadı", 500
            
        try:
            parts = month.split("-")
            year_val = int(parts[0])
            month_val = int(parts[1])
            normalized_month = f"{year_val}-{str(month_val).zfill(2)}"
        except:
            normalized_month = month
            year_val, month_val = map(int, datetime.now().strftime("%Y-%m").split("-"))

        employees = conn.execute("SELECT id, name FROM employees ORDER BY name ASC").fetchall()
        
        wb = Workbook()
        ws = wb.active
        ws.title = f"Puantaj-{normalized_month}"
        ws.views.sheetView[0].showGridLines = True

        main_header_font = Font(bold=True, color="0062FF", size=14, name="Arial")
        sub_header_font = Font(bold=True, color="1A1A1A", size=11, name="Arial")
        info_label_font = Font(bold=True, color="555555", size=9, name="Arial")
        info_val_font = Font(bold=True, color="0062FF", size=9.5, name="Arial")
        
        header_font = Font(bold=True, color="FFFFFF", size=8.5, name="Arial")
        cell_font = Font(size=8.5, name="Arial")
        bold_cell_font = Font(bold=True, size=8.5, name="Arial")
        blue_label_font = Font(bold=True, color="0062FF", size=8.5, name="Arial")
        
        black_fill = PatternFill(start_color="1A1A1A", end_color="1A1A1A", fill_type="solid")     
        grey_fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")      
        weekend_fill = PatternFill(start_color="EAEAEA", end_color="EAEAEA", fill_type="solid")   
        hzm_blue_fill = PatternFill(start_color="0062FF", end_color="0062FF", fill_type="solid")  
        
        thin_border = Border(
            left=Side(style='thin', color='D9D9D9'), right=Side(style='thin', color='D9D9D9'), 
            top=Side(style='thin', color='D9D9D9'), bottom=Side(style='thin', color='D9D9D9')
        )
        
        center_align = Alignment(horizontal="center", vertical="center")
        left_align = Alignment(horizontal="left", vertical="center", indent=1)

        ws.row_dimensions[2].height = 26
        ws.merge_cells('A2:H2')
        title_top = ws.cell(row=2, column=1, value="TSİ INTEGRATION SYSTEM")
        title_top.font = main_header_font
        title_top.alignment = left_align
        
        ws.row_dimensions[4].height = 20
        ws.row_dimensions[5].height = 20
        
        ws.merge_cells('A4:B4')   
        ws.merge_cells('C4:F4')   
        ws.merge_cells('G4:J4')   
        ws.merge_cells('K4:W4')   
        
        lbl1 = ws.cell(row=4, column=1, value="DÖNEM / AY:")
        val1 = ws.cell(row=4, column=3, value=normalized_month)
        lbl2 = ws.cell(row=4, column=7, value="DÖKÜMAN TİPİ:")
        val2 = ws.cell(row=4, column=11, value="AYLIK PUANTAJ RAPORU")
        
        ws.merge_cells('A5:B5')   
        ws.merge_cells('C5:M5')   
        ws.merge_cells('N5:Q5')   
        ws.merge_cells('R5:W5')   
        
        lbl3 = ws.cell(row=5, column=1, value="SİSTEM:")
        val3 = ws.cell(row=5, column=3, value="TECHNICAL STATISTICS & INTEGRATION")
        lbl4 = ws.cell(row=5, column=14, value="RAPOR TARİHİ:")
        val4 = ws.cell(row=5, column=18, value=datetime.now().strftime("%d.%m.%Y"))
        
        for lbl_cell in [lbl1, lbl2, lbl3, lbl4]:
            lbl_cell.font = info_label_font
            lbl_cell.alignment = left_align
            
        for val_cell in [val1, val2, val3, val4]:
            val_cell.font = info_val_font
            val_cell.alignment = left_align

        weekend_cols = []
        for day in range(1, 32):
            try:
                d_obj = date(year_val, month_val, day)
                if d_obj.weekday() in [5, 6]: 
                    weekend_cols.append(day + 1)
            except ValueError:
                pass 

        attendance_map = {}
        if employees:
            all_data = conn.execute("""
                SELECT employee_id, date, work_value, COALESCE(extra_hours, 0) as extra_hours 
                FROM attendance 
                WHERE strftime('%Y-%m', date) = ?
            """, (normalized_month,)).fetchall()
            
            for row in all_data:
                try:
                    emp_id_key = int(row['employee_id'])
                    date_key = row['date']
                    w_val = row['work_value']
                    e_hrs = row['extra_hours']
                except:
                    emp_id_key = int(row[0])
                    date_key = row[1]
                    w_val = row[2]
                    e_hrs = row[3]
                    
                if emp_id_key not in attendance_map:
                    attendance_map[emp_id_key] = {}
                attendance_map[emp_id_key][date_key] = {'work': w_val, 'extra': e_hrs}

        t1_title_row = 7
        ws.cell(row=t1_title_row, column=1, value="1. ANA PUANTAJ MATRİSİ").font = sub_header_font
        
        t1_start_row = 8
        headers_t1 = ["PERSONEL ADI"]
        for day in range(1, 32): 
            headers_t1.append(str(day))
        headers_t1.append("TOPLAM GÜN")
        
        ws.row_dimensions[t1_start_row].height = 24
        for col_idx, text in enumerate(headers_t1, start=1):
            cell = ws.cell(row=t1_start_row, column=col_idx, value=text)
            cell.font = header_font
            cell.fill = black_fill
            cell.alignment = center_align if col_idx > 1 else left_align
            cell.border = thin_border

        current_row = t1_start_row + 1
        
        for emp in employees:
            try:
                emp_id = int(emp['id'])
                emp_name = emp['name']
            except:
                emp_id = int(emp[0])
                emp_name = emp[1]

            ws.row_dimensions[current_row].height = 22
            n_cell = ws.cell(row=current_row, column=1, value=str(emp_name).upper())
            n_cell.font = cell_font
            n_cell.border = thin_border
            n_cell.alignment = left_align
            
            for day in range(1, 32):
                col_idx = day + 1
                date_str = f"{normalized_month}-{str(day).zfill(2)}"
                
                work_val = 0.0
                if emp_id in attendance_map and date_str in attendance_map[emp_id]:
                    work_val = float(attendance_map[emp_id][date_str]['work'] or 0.0)

                cell_final_value = work_val if work_val > 0 else "-"
                
                day_cell = ws.cell(row=current_row, column=col_idx, value=cell_final_value)
                day_cell.font = cell_font
                day_cell.border = thin_border
                day_cell.alignment = center_align
                if isinstance(cell_final_value, (int, float)):
                    day_cell.number_format = '0.0'
                if col_idx in weekend_cols:
                    day_cell.fill = weekend_fill
            
            total_day_cell = ws.cell(row=current_row, column=33, value=f"=SUM(B{current_row}:AF{current_row})")
            total_day_cell.font = bold_cell_font
            total_day_cell.border = thin_border
            total_day_cell.alignment = center_align
            total_day_cell.number_format = '#,##0.0'
            
            current_row += 1

        last_row_t1 = current_row
        ws.row_dimensions[last_row_t1].height = 24
        
        footer_label_t1 = ws.cell(row=last_row_t1, column=1, value="GÜNLÜK TOPLAM")
        footer_label_t1.font = blue_label_font
        footer_label_t1.fill = grey_fill
        footer_label_t1.alignment = left_align
        footer_label_t1.border = thin_border

        for i in range(2, 33):
            col_letter = get_column_letter(i)
            cell = ws.cell(row=last_row_t1, column=i, value=f"=SUM({col_letter}{t1_start_row+1}:{col_letter}{last_row_t1-1})")
            cell.font = bold_cell_font
            cell.fill = grey_fill
            cell.alignment = center_align
            cell.border = thin_border
            cell.number_format = '#,##0.0'
            if i in weekend_cols:
                cell.fill = weekend_fill

        g_total_cell_t1 = ws.cell(row=last_row_t1, column=33, value=f"=SUM(AG{t1_start_row+1}:AG{last_row_t1-1})")
        g_total_cell_t1.font = Font(bold=True, color="FFFFFF", size=8.5, name="Arial")
        g_total_cell_t1.fill = hzm_blue_fill
        g_total_cell_t1.alignment = center_align
        g_total_cell_t1.border = thin_border
        g_total_cell_t1.number_format = '#,##0.0'

        t2_title_row = last_row_t1 + 3
        ws.cell(row=t2_title_row, column=1, value="2. AYLIK TOPLAM MESAİ RAPORU").font = sub_header_font
        
        t2_start_row = t2_title_row + 1
        ws.row_dimensions[t2_start_row].height = 24
        
        ws.merge_cells(start_row=t2_start_row, start_column=1, end_row=t2_start_row, end_column=8)   
        ws.merge_cells(start_row=t2_start_row, start_column=9, end_row=t2_start_row, end_column=20)  
        ws.merge_cells(start_row=t2_start_row, start_column=21, end_row=t2_start_row, end_column=33) 
        
        h_cell1 = ws.cell(row=t2_start_row, column=1, value="PERSONEL ADI")
        h_cell2 = ws.cell(row=t2_start_row, column=9, value="TOPLAM MESAİ SAATİ")
        h_cell3 = ws.cell(row=t2_start_row, column=21, value="GÜN KARŞILIĞI (SAAT / 8)")
        
        for c_cell, is_left in [(h_cell1, True), (h_cell2, False), (h_cell3, False)]:
            c_cell.font = header_font
            c_cell.fill = black_fill
            c_cell.alignment = left_align if is_left else center_align
            
        for col in range(1, 34):
            ws.cell(row=t2_start_row, column=col).border = thin_border

        current_row_t2 = t2_start_row + 1
        
        for emp in employees:
            try:
                emp_id = int(emp['id'])
                emp_name = emp['name']
            except:
                emp_id = int(emp[0])
                emp_name = emp[1]

            ws.row_dimensions[current_row_t2].height = 22
            
            ws.merge_cells(start_row=current_row_t2, start_column=1, end_row=current_row_t2, end_column=8)
            ws.merge_cells(start_row=current_row_t2, start_column=9, end_row=current_row_t2, end_column=20)
            ws.merge_cells(start_row=current_row_t2, start_column=21, end_row=current_row_t2, end_column=33)
            
            n_cell_t2 = ws.cell(row=current_row_t2, column=1, value=str(emp_name).upper())
            n_cell_t2.font = cell_font
            n_cell_t2.alignment = left_align
            
            total_extra_hours = 0.0
            if emp_id in attendance_map:
                total_extra_hours = sum(float(day_data['extra'] or 0.0) for day_data in attendance_map[emp_id].values())

            hours_cell = ws.cell(row=current_row_t2, column=9, value=total_extra_hours if total_extra_hours > 0 else 0)
            hours_cell.font = cell_font
            hours_cell.alignment = center_align
            hours_cell.number_format = '#,##0.0'
            
            converted_days_cell = ws.cell(row=current_row_t2, column=21, value=f"=I{current_row_t2}/8.0")
            converted_days_cell.font = bold_cell_font
            converted_days_cell.alignment = center_align
            converted_days_cell.number_format = '#,##0.00'
            
            for col in range(1, 34):
                ws.cell(row=current_row_t2, column=col).border = thin_border
                
            current_row_t2 += 1

        last_row_t2 = current_row_t2
        ws.row_dimensions[last_row_t2].height = 24
        
        ws.merge_cells(start_row=last_row_t2, start_column=1, end_row=last_row_t2, end_column=8)
        ws.merge_cells(start_row=last_row_t2, start_column=9, end_row=last_row_t2, end_column=20)
        ws.merge_cells(start_row=last_row_t2, start_column=21, end_row=last_row_t2, end_column=33)
        
        footer_label_t2 = ws.cell(row=last_row_t2, column=1, value="GENEL TOPLAM")
        footer_label_t2.font = blue_label_font
        footer_label_t2.fill = grey_fill
        footer_label_t2.alignment = left_align

        g_hours_cell = ws.cell(row=last_row_t2, column=9, value=f"=SUM(I{t2_start_row+1}:I{last_row_t2-1})")
        g_hours_cell.font = Font(bold=True, color="FFFFFF", size=8.5, name="Arial")
        g_hours_cell.fill = hzm_blue_fill
        g_hours_cell.alignment = center_align
        g_hours_cell.number_format = '#,##0.0'

        g_converted_days_cell = ws.cell(row=last_row_t2, column=21, value=f"=SUM(U{t2_start_row+1}:U{last_row_t2-1})")
        g_converted_days_cell.font = Font(bold=True, color="FFFFFF", size=8.5, name="Arial")
        g_converted_days_cell.fill = hzm_blue_fill
        g_converted_days_cell.alignment = center_align
        g_converted_days_cell.number_format = '#,##0.00'

        for col in range(1, 34):
            c_bt = ws.cell(row=last_row_t2, column=col)
            c_bt.border = thin_border
            if col <= 8:
                c_bt.fill = grey_fill

        ws.column_dimensions['A'].width = 24 
        for i in range(2, 33):
            ws.column_dimensions[get_column_letter(i)].width = 4.8
        ws.column_dimensions['AG'].width = 13

        from PIL import Image as PILImage
        
        text_row = last_row_t2 + 2
        ws.row_dimensions[text_row].height = 24
        ws.merge_cells(start_row=text_row, start_column=1, end_row=text_row, end_column=33)
        
        mühür_cell = ws.cell(row=text_row, column=1, value="Bu Belge HZM SOFTWARE & DESIGN SERVICES Tarafından Geliştirilen TSI: Technical Statistics & Integration Sistemi İle Hazırlanmıştır.")
        mühür_cell.font = Font(size=9, color="333333", italic=True, name="Arial")
        mühür_cell.alignment = center_align
        
        logo_row = text_row + 2
        target_height_px = 50  

        logo_path = 'static/img/logo.png'
        logo_path2 = 'static/img/TSI_logo.png'
        
        if os.path.exists(logo_path):
            with PILImage.open(logo_path) as img1:
                orig_w, orig_h = img1.size
            scale_left = target_height_px / orig_h
            
            img_left = OpenpyxlImage(logo_path)
            img_left.width = int(orig_w * scale_left)
            img_left.height = target_height_px
            ws.add_image(img_left, f'A{logo_row}')
            
        if os.path.exists(logo_path2):
            with PILImage.open(logo_path2) as img2:
                orig_w2, orig_h2 = img2.size
            scale_right = target_height_px / orig_h2
            
            img_right = OpenpyxlImage(logo_path2)
            img_right.width = int(orig_w2 * scale_right)
            img_right.height = target_height_px
            ws.add_image(img_right, f'AG{logo_row}')
            
        base_name = f"{normalized_month}_Puantaj_Raporu"
        extension = ".xlsx"
        file_name = f"{base_name}{extension}"
        downloads_path = str(Path.home() / "Downloads")
        full_path = os.path.join(downloads_path, file_name)
        
        counter = 1
        while True:
            if os.path.exists(full_path):
                try:
                    with open(full_path, 'r+'):
                        pass
                    break
                except IOError:
                    file_name = f"{base_name} ({counter}){extension}"
                    full_path = os.path.join(downloads_path, file_name)
                    counter += 1
            else:
                break

        wb.save(full_path)
        conn.close()

        if os.name == 'nt': 
            os.startfile(downloads_path)
            
        return send_file(full_path, as_attachment=True)

    except Exception as e:
        print(f"TSİ Orijinal Puantaj Hatası: {e}")
        return f"Sistem Hatası: {str(e)}", 500

@app.route("/export-detay-excel/<int:emp_id>")
def export_detay_excel(emp_id):
    start_date = request.args.get('start_date') 
    end_date = request.args.get('end_date')

    conn = get_conn()
    try:
        emp_info = conn.execute("SELECT name, daily_wage FROM employees WHERE id = ?", (emp_id,)).fetchone()
        emp_name = emp_info['name'] if emp_info else f"Personel_{emp_id}"
        current_daily_wage = emp_info['daily_wage'] if emp_info else 0
        
        query_attendance = """
            SELECT 
                SUM(work_value) as toplam_normal_gun,
                SUM(COALESCE(extra_hours, 0)) as toplam_mesai_saat,
                SUM(
                    (work_value + (COALESCE(extra_hours, 0) / 8.0)) * CASE WHEN daily_wage > 0 THEN daily_wage ELSE ? END
                ) as toplam_hakedis
            FROM attendance 
            WHERE employee_id = ?
        """
        params_att = [current_daily_wage, emp_id]

        if start_date and end_date:
            query_attendance += " AND date BETWEEN ? AND ?"
            params_att.extend([start_date, end_date])

        stats = conn.execute(query_attendance, params_att).fetchone()

        query_asgari = "SELECT SUM(amount) FROM payments WHERE employee_id = ? AND description = 'Ödeme 1 (Maaş)'"
        query_avans = "SELECT SUM(amount) FROM payments WHERE employee_id = ? AND (description IS NULL OR description != 'Ödeme 1 (Maaş)')"
        params_pay = [emp_id]

        if start_date and end_date:
            query_asgari += " AND date BETWEEN ? AND ?"
            query_avans += " AND date BETWEEN ? AND ?"
            params_pay.extend([start_date, end_date])

        asgari = conn.execute(query_asgari, params_pay if not (start_date and end_date) else [emp_id, start_date, end_date]).fetchone()[0] or 0
        avans = conn.execute(query_avans, params_pay if not (start_date and end_date) else [emp_id, start_date, end_date]).fetchone()[0] or 0

        normal_gun = stats['toplam_normal_gun'] or 0
        mesai_saat = stats['toplam_mesai_saat'] or 0
        mesai_gun_karsiligi = mesai_saat / 8.0 
        
        yevmiye_tl = stats['toplam_hakedis'] or 0
        net_kalan = yevmiye_tl - (asgari + avans)

        grand_total_days = normal_gun + mesai_gun_karsiligi
        avg_wage = (yevmiye_tl / grand_total_days) if grand_total_days > 0 else current_daily_wage

        clean_name = emp_name.replace(" ", "_")
        date_str = f"_{start_date}_to_{end_date}" if start_date else ""
        file_name = f"TSI_{clean_name}{date_str}_Detay_Rapor.xlsx"
        downloads_path = str(Path.home() / "Downloads")
        final_path = os.path.join(downloads_path, file_name)

        workbook = xlsxwriter.Workbook(final_path)
        worksheet = workbook.add_worksheet('OZET')

        title_fmt = workbook.add_format({'bold': True, 'size': 14, 'align': 'center', 'bg_color': '#000000', 'font_color': '#0dcaf0'})
        info_fmt  = workbook.add_format({'bold': True, 'size': 10, 'font_color': '#666666'})
        h_style   = workbook.add_format({'bold': True, 'bg_color': '#1a1a1a', 'font_color': '#d1d1d1', 'border': 1, 'align': 'center'})
        n_style   = workbook.add_format({'border': 1, 'align': 'center', 'num_format': '#,##0.00 "₺"'})
        g_style   = workbook.add_format({'border': 1, 'align': 'center', 'bold': True})
        
        worksheet.merge_range('A1:H1', 'HZM TSI - PERSONEL ÖZET RAPORU', title_fmt)
        worksheet.write('A3', 'PERSONEL:', info_fmt)
        worksheet.write('B3', emp_name.upper())
        worksheet.write('A4', 'ARALIK:', info_fmt)
        worksheet.write('B4', f"{start_date} / {end_date}" if start_date else "TÜM ZAMANLAR")

        headers = [
            "TOPLAM MESAİ (GÜN)", 
            "MESAİ SAAT", 
            "MESAİ GÜN KARŞILIĞI", 
            "ORT. YEVMİYE", 
            "BRÜT HAKEDİŞ", 
            "ASGARİ ÖDEME", 
            "AVANS", 
            "NET BAKİYE"
        ]
        for i, h in enumerate(headers): 
            worksheet.write(6, i, h, h_style)

        worksheet.write(7, 0, f"{normal_gun:.2f} G", g_style)
        worksheet.write(7, 1, f"{mesai_saat:.2f} Sa", g_style)
        worksheet.write(7, 2, f"{mesai_gun_karsiligi:.2f} G", g_style)
        worksheet.write(7, 3, avg_wage, n_style)
        worksheet.write(7, 4, yevmiye_tl, n_style)
        worksheet.write(7, 5, asgari, n_style)
        worksheet.write(7, 6, avans, n_style)
        
        kalan_color = '#28A745' if net_kalan >= 0 else '#DC3545'
        k_style = workbook.add_format({'border': 1, 'align': 'center', 'bold': True, 'num_format': '#,##0.00 "₺"', 'font_color': kalan_color})
        worksheet.write(7, 7, net_kalan, k_style)

        from PIL import Image as PILImage

        normal_text_fmt = workbook.add_format({'font_size': 9.5, 'font_color': '#333333', 'align': 'center', 'valign': 'vcenter', 'text_wrap': True})
        bold_text_fmt   = workbook.add_format({'font_size': 9.5, 'font_color': '#000000', 'bold': True, 'align': 'center', 'valign': 'vcenter', 'text_wrap': True})
        
        worksheet.merge_range('B11:G11', "", normal_text_fmt)
        worksheet.write_rich_string('B11',
            'Bu Belge ', 
            bold_text_fmt, 'HZM SOFTWARE & DESIGN SERVICES', 
            normal_text_fmt, ' Tarafından Geliştirilen ', 
            bold_text_fmt, 'TSI: Technical Statistics & Integration', 
            normal_text_fmt, ' Sistemi İle Hazırlanmıştır.',
            normal_text_fmt
        )
        
        target_height_px = 45
        worksheet.set_row(10, 42)  

        logo_path = 'static/img/logo.png'
        logo_path2 = 'static/img/TSI_logo.png'

        if os.path.exists(logo_path):
            with PILImage.open(logo_path) as img1:
                orig_w, orig_h = img1.size
            scale_left = target_height_px / orig_h
            
            worksheet.insert_image('A11', logo_path, {
                'x_scale': scale_left, 
                'y_scale': scale_left, 
                'object_position': 1,
                'x_offset': 10,  
                'y_offset': 6    
            })
            
        if os.path.exists(logo_path2):
            with PILImage.open(logo_path2) as img2:
                orig_w2, orig_h2 = img2.size
            scale_right = target_height_px / orig_h2
            
            worksheet.insert_image('H11', logo_path2, {
                'x_scale': scale_right, 
                'y_scale': scale_right, 
                'object_position': 1,
                'x_offset': 15,  
                'y_offset': 6
            })

        worksheet.set_column('A:H', 20)
        workbook.close()

        if os.name == 'nt': 
            os.startfile(downloads_path)

        return send_file(final_path, as_attachment=True)

    except Exception as e:
        print(f"Excel Hatası: {e}")
        return f"Hata: {str(e)}", 500
    finally:
        conn.close()

@app.route("/logout-company")
def logout_company():
    session.clear()
    return redirect(url_for('select_company'))

# =================================================================
# ÇALIŞTIRMA
# =================================================================
TSI_PRIVATE_PORT = 58432

if __name__ == "__main__":
    threading.Thread(
        target=lambda: serve(app, host="127.0.0.1", port=TSI_PRIVATE_PORT, threads=12), 
        daemon=True
    ).start()

    import time
    time.sleep(0.5)
    
    webview.create_window(
        'TSİ Teknik İstatistik & Entegrasyon Yönetim Paneli', 
        f'http://127.0.0.1:{TSI_PRIVATE_PORT}',
        maximized=True
    )
    webview.start()